import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# 새로운 엑셀파일 생성
total_wb = Workbook()

# 현재 활성화된 시트 선택
total_ws = total_wb.active

# 시트 이름 변경

total_ws.title = "data"

# 헤더 추가
total_ws.append(['순번','제품명','가격','수량','합계'])

# 데이터 파일 갯수 확인 후 리스트에 담기
import glob
file_list=[]

# glob.glob는 파일 시스템에서 특정 패턴에 맞는 파일명들을 찾을 때 사용
file_list=glob.glob("예시파일/*.xlsx")

# 이제 순번을 다시 정한 다음 저장하기
for excel_name in file_list: 
    wb = load_workbook(filename = excel_name, data_only='True')
    ws = wb.active

    # iter_rows는  워크시트의 모든 행을 순회하는 이터레이터(iterator)를 반환
    for row in ws.iter_rows(min_row=2): # 첫 번째 행을 제외하고 모든 행을 순회
        data = []
        for cell in row:
            data.append(cell.value)
        total_ws.append(data)

# total_ws 워크시트의 모든 행을 순회하여 순번을 부여
for row in total_ws.iter_rows(min_row=2, max_col=1):
    for cell in row:
        cell.value = row[0].row -1  #순번은 '행번호-1'로 부여
# Alignment 객체 생성 (가로 및 세로 중앙 정렬)
alignment = Alignment(horizontal='center', vertical='center')

# Font 객체 생성 (폰트 스타일 설정)
font = Font(name='Calibri', size=12, bold=True)

# Border 객체 생성 (테두리 스타일 설정)
side = Side(border_style="thin", color="7634c7")  # 테두리 스타일 지정
border = Border(left=side, right=side, top=side, bottom=side)

# 1행의 모든 셀에 스타일 적용
for cell in total_ws['1']:
    cell.alignment = alignment
    cell.font = font
    cell.border = border

# 변경 사항 저장
total_wb.save('styled_workbook2.xlsx')


